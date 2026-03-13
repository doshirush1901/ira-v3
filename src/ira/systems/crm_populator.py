"""CRM population pipeline — classify and import contacts.

Extracts contacts from Gmail inbox, the Qdrant knowledge base, and
Neo4j, classifies each via Delphi (with Clio-style cross-referencing
for evidence), and inserts only CRM-eligible contacts (live customer,
past customer, lead with interactions, lead without interactions).

Usage::

    ira populate-crm [--dry-run] [--source all|gmail|kb|neo4j]
"""

from __future__ import annotations

import asyncio
import csv
import json
import logging
import mailbox
import re as _re_module
from email.utils import getaddresses, parseaddr
from pathlib import Path
from typing import Any

import httpx

from ira.agents.delphi import Delphi
from ira.config import get_settings
from ira.data.crm import CRMDatabase
from ira.data.models import ContactType, DealStage

logger = logging.getLogger(__name__)

# Base path for imports (repo root = parents[3] from src/ira/systems/crm_populator.py)
_REPO_ROOT = Path(__file__).resolve().parents[3]
_IMPORTS_LEADS_DIR = _REPO_ROOT / "data" / "imports" / "24_WebSite_Leads"
_IMPORTS_07_DIR = _REPO_ROOT / "data" / "imports" / "07_Leads_and_Contacts"
_IMPORTS_ROOT = _REPO_ROOT / "data" / "imports"
_TAKEOUT_INGEST_DIR = _REPO_ROOT / "data" / "takeout_ingest"
_MAX_TAKEOUT_SENT_CONTACTS = 2000

_VALID_CRM_TYPES = {
    "LIVE_CUSTOMER": ContactType.LIVE_CUSTOMER,
    "PAST_CUSTOMER": ContactType.PAST_CUSTOMER,
    "LEAD_WITH_INTERACTIONS": ContactType.LEAD_WITH_INTERACTIONS,
    "LEAD_NO_INTERACTIONS": ContactType.LEAD_NO_INTERACTIONS,
}

_OWN_DOMAINS = frozenset({"machinecraft.org", "machinecraft.in"})

_REJECT_PREFIXES = frozenset({
    "noreply", "no-reply", "donotreply", "do-not-reply", "billing",
    "alerts", "newsletter", "news", "notifications", "notify",
    "mailer", "marketing", "promo", "info@members", "hello@mail.",
    "hello@email.", "hello@e.", "share@email.", "team@mail.",
    "stories-recap", "support@mail.", "info@email.",
})

_REJECT_DOMAINS = frozenset({
    "instagram.com", "facebook.com", "twitter.com", "linkedin.com",
    "netflix.com", "spotify.com", "amazon.com", "google.com",
    "apple.com", "microsoft.com", "github.com", "openai.com",
    "moneycontrol.com", "squarespace.com", "medium.com",
    "substack.com", "mailchimp.com", "hubspot.com",
    "singaporeair.com", "shakeshack.com", "avg.com", "cursor.com",
    "railway.app", "sampark.gov.in", "yatramailers.com",
    "sunglasshut.com", "marcjacobs.com", "nordiskemedier.dk",
    "economictimesnews.com", "fairygodboss.com", "careem.com",
    "impactguru.com", "thomasnet.com", "freshplaza.com",
    "nin-nin.fr", "sarahssilks.com", "rforrabbit.com",
    "bombaysweetshop.com", "littletikes.com", "ethoswatches.com",
    "nevertoosmall.com", "zamaorganics.com", "mapmygenome.in",
    "cgboost.com", "strategyzer.com", "camsonline.com",
    "smarttouchswitch.io", "onepercentclub.io", "dsij.in",
    "msmegrowthhub.com", "danmartell.com",
})


def _is_obviously_not_business(email: str) -> bool:
    """Fast pre-filter to reject consumer newsletters and automated senders."""
    local, _, domain = email.partition("@")
    if not domain:
        return True

    if domain in _REJECT_DOMAINS:
        return True

    parent_domain = ".".join(domain.split(".")[-2:])
    if parent_domain in _REJECT_DOMAINS:
        return True

    for prefix in _REJECT_PREFIXES:
        if email.startswith(prefix):
            return True

    return False


class CRMPopulator:
    """Orchestrates contact extraction, classification, and CRM insertion."""

    def __init__(
        self,
        delphi: Delphi,
        crm: CRMDatabase,
        *,
        dry_run: bool = False,
        skip_classification: bool = False,
        event_bus: Any | None = None,
    ) -> None:
        self._delphi = delphi
        self._crm = crm
        self._dry_run = dry_run
        self._skip_classification = skip_classification
        self._event_bus = event_bus

        settings = get_settings()
        self._qdrant_url = settings.qdrant.url
        self._qdrant_api_key = settings.qdrant.api_key.get_secret_value()
        self._qdrant_collection = settings.qdrant.collection

        self._stats: dict[str, int] = {
            "total_extracted": 0,
            "classified": 0,
            "inserted": 0,
            "skipped_duplicate": 0,
            "skipped_rejected": 0,
            "skipped_no_email": 0,
            "errors": 0,
        }
        self._classifications: list[dict[str, Any]] = []

    @property
    def stats(self) -> dict[str, int]:
        return dict(self._stats)

    @property
    def classifications(self) -> list[dict[str, Any]]:
        return list(self._classifications)

    async def populate(
        self,
        sources: list[str] | None = None,
        *,
        after: str = "",
        before: str = "",
    ) -> dict[str, Any]:
        """Run the full population pipeline.

        Parameters
        ----------
        sources : list[str] | None
            Data sources to use (``gmail``, ``kb``, ``neo4j``).
            Defaults to all three.
        after, before : str
            Date range for Gmail extraction in ``YYYY/MM/DD`` format.
        """
        use = set(sources) if sources else {"gmail", "kb", "neo4j", "imports", "imports_07", "takeout_sent"}

        contacts: list[dict[str, Any]] = []

        if "gmail" in use:
            contacts.extend(await self._extract_from_gmail(after=after, before=before))
        if "kb" in use:
            contacts.extend(await self._extract_from_kb())
        if "neo4j" in use:
            contacts.extend(await self._extract_from_neo4j())
        if "imports" in use:
            contacts.extend(await self._extract_from_imports())
        if "imports_07" in use:
            contacts.extend(await self._extract_from_imports_07())
        if "takeout_sent" in use:
            contacts.extend(await self._extract_from_takeout_sent())

        contacts = self._deduplicate(contacts)
        self._stats["total_extracted"] = len(contacts)
        logger.info("Extracted %d unique contacts", len(contacts))

        for i, contact in enumerate(contacts):
            try:
                await self._process_contact(contact)
                if (i + 1) % 10 == 0:
                    logger.info(
                        "Progress: %d/%d (inserted=%d, rejected=%d)",
                        i + 1, len(contacts),
                        self._stats["inserted"],
                        self._stats["skipped_rejected"],
                    )
            except Exception:
                self._stats["errors"] += 1
                logger.exception(
                    "Failed to process contact: %s",
                    contact.get("email", "unknown"),
                )

        logger.info("CRM population complete: %s", self._stats)
        return {
            "stats": self._stats,
            "dry_run": self._dry_run,
            "classifications": self._classifications,
        }

    # ── Processing ───────────────────────────────────────────────────────

    async def _process_contact(self, contact: dict[str, Any]) -> None:
        email = contact.get("email", "").strip().lower()
        if not email or "@" not in email:
            self._stats["skipped_no_email"] += 1
            return

        domain = email.split("@", 1)[1]
        if domain in _OWN_DOMAINS:
            self._stats["skipped_rejected"] += 1
            self._classifications.append({
                "email": email, "company": contact.get("company", ""),
                "contact_type": "OWN_COMPANY", "confidence": "HIGH",
                "reasoning": "Internal Machinecraft domain",
            })
            return

        if _is_obviously_not_business(email):
            self._stats["skipped_rejected"] += 1
            self._classifications.append({
                "email": email, "company": contact.get("company", ""),
                "contact_type": "OTHER", "confidence": "HIGH",
                "reasoning": "Consumer/newsletter/automated sender (pre-filter)",
            })
            return

        existing = await self._crm.get_contact_by_email(email)
        if existing is not None:
            self._stats["skipped_duplicate"] += 1
            # Backfill deal if we have machine_model from imports and contact has no deal
            machine_model = contact.get("machine_model")
            if machine_model and not self._dry_run:
                try:
                    deals = await self._crm.list_deals(filters={"contact_id": str(existing.id)})
                    if not deals:
                        await self._crm.create_deal(
                            contact_id=str(existing.id),
                            title=contact.get("company") or f"Import {email}",
                            value=0,
                            stage=DealStage.CONTACTED,
                            machine_model=machine_model,
                        )
                        logger.info("Backfilled deal for %s (%s)", email, machine_model)
                except Exception:
                    logger.debug("Failed to backfill deal for %s", email, exc_info=True)
            return

        contact_type_hint = contact.get("contact_type_hint")
        if contact_type_hint and contact_type_hint in _VALID_CRM_TYPES:
            classification = {
                "contact_type": contact_type_hint,
                "confidence": "HIGH",
                "reasoning": "From imports lead context file",
            }
            self._stats["classified"] += 1
            self._classifications.append({
                "email": email,
                "name": contact.get("name", ""),
                "company": contact.get("company", ""),
                **classification,
            })
            contact_type_str = contact_type_hint
        elif self._skip_classification:
            classification = {
                "contact_type": "LEAD_NO_INTERACTIONS",
                "confidence": "LOW",
                "reasoning": "Skipped (--no-classify); Qdrant/KB not required.",
            }
            self._stats["classified"] += 1
            self._classifications.append({
                "email": email,
                "name": contact.get("name", ""),
                "company": contact.get("company", ""),
                **classification,
            })
            contact_type_str = "LEAD_NO_INTERACTIONS"
        else:
            evidence = await self._gather_evidence(contact)
            contact["order_history"] = evidence.get("order_history", "")
            contact["email_evidence"] = evidence.get("email_evidence", "")

            classification = await self._classify(contact)
            self._stats["classified"] += 1

            self._classifications.append({
                "email": email,
                "name": contact.get("name", ""),
                "company": contact.get("company", ""),
                **classification,
            })

            contact_type_str = classification.get("contact_type", "")
        if contact_type_str not in _VALID_CRM_TYPES:
            self._stats["skipped_rejected"] += 1
            logger.debug(
                "Rejected %s (%s): %s — %s",
                email, contact.get("company", ""),
                contact_type_str, classification.get("reasoning", ""),
            )
            return

        if self._dry_run:
            self._stats["inserted"] += 1
            logger.info(
                "[DRY RUN] Would insert: %s (%s) as %s [%s]",
                email, contact.get("company", ""),
                contact_type_str, classification.get("confidence", "?"),
            )
            return

        company_name = contact.get("company", "")
        company_id = None
        if company_name:
            company_id = await self._ensure_company(company_name, contact.get("region"))

        created_contact = await self._crm.create_contact(
            name=contact.get("name", email.split("@")[0]),
            email=email,
            company_id=company_id,
            role=contact.get("role"),
            phone=contact.get("phone"),
            source=contact.get("source", "populator"),
            contact_type=_VALID_CRM_TYPES[contact_type_str],
            lead_score=0.0,
        )
        self._stats["inserted"] += 1
        logger.info("Inserted: %s (%s) as %s", email, company_name, contact_type_str)

        machine_model = contact.get("machine_model")
        if machine_model and not self._dry_run:
            try:
                await self._crm.create_deal(
                    contact_id=str(created_contact.id),
                    title=company_name or f"Import {email}",
                    value=0,
                    stage=DealStage.CONTACTED,
                    machine_model=machine_model,
                )
            except Exception:
                logger.debug("Failed to create deal for %s", email, exc_info=True)

        if self._event_bus is not None:
            from ira.systems.data_event_bus import DataEvent, EventType, SourceStore
            try:
                await self._event_bus.emit(DataEvent(
                    event_type=EventType.CONTACT_CLASSIFIED,
                    entity_type="contact",
                    entity_id=email,
                    payload={
                        "email": email,
                        "name": contact.get("name", ""),
                        "company": company_name,
                        "contact_type": contact_type_str,
                        "confidence": classification.get("confidence", ""),
                        "reasoning": classification.get("reasoning", ""),
                        "source": contact.get("source", "populator"),
                    },
                    source_store=SourceStore.POPULATOR,
                ))
            except Exception:
                logger.warning("Populator event emission failed", exc_info=True)

    # ── Evidence gathering (Clio cross-referencing) ──────────────────────

    async def _gather_evidence(self, contact: dict[str, Any]) -> dict[str, Any]:
        """Search the KB for order/quote documents mentioning this contact's company."""
        company = contact.get("company", "")
        name = contact.get("name", "")

        if not company and not name:
            return {}

        search_term = company or name
        evidence: dict[str, Any] = {}

        try:
            headers = {"Content-Type": "application/json"}
            if self._qdrant_api_key:
                headers["api-key"] = self._qdrant_api_key

            async with httpx.AsyncClient(timeout=10) as client:
                resp = await client.post(
                    f"{self._qdrant_url}/collections/{self._qdrant_collection}/points/scroll",
                    headers=headers,
                    json={
                        "limit": 200,
                        "with_payload": {"include": ["content", "source", "source_category"]},
                        "with_vector": False,
                    },
                )
                resp.raise_for_status()
                points = resp.json()["result"]["points"]

            order_hits: list[str] = []
            quote_hits: list[str] = []
            search_lower = search_term.lower()

            for pt in points:
                content = (pt["payload"].get("content") or "").lower()
                source = (pt["payload"].get("source") or "").lower()
                cat = (pt["payload"].get("source_category") or "").lower()

                if search_lower not in content and search_lower not in source:
                    continue

                if any(kw in cat for kw in ("order", "po", "purchase")):
                    order_hits.append(pt["payload"].get("content", "")[:300])
                elif any(kw in cat for kw in ("quote", "offer", "pricing", "proposal")):
                    quote_hits.append(pt["payload"].get("content", "")[:300])
                elif any(kw in content for kw in ("purchase order", "po number", "order confirmation", "delivery")):
                    order_hits.append(pt["payload"].get("content", "")[:300])
                elif any(kw in content for kw in ("quote", "quotation", "offer", "proposal")):
                    quote_hits.append(pt["payload"].get("content", "")[:300])

            if order_hits:
                evidence["order_history"] = f"Found {len(order_hits)} order/PO documents:\n" + "\n".join(
                    f"  - {h[:200]}" for h in order_hits[:3]
                )
            if quote_hits:
                existing = evidence.get("order_history", "")
                evidence["order_history"] = (
                    existing + f"\nFound {len(quote_hits)} quote documents:\n" + "\n".join(
                        f"  - {h[:200]}" for h in quote_hits[:3]
                    )
                ).strip()

        except Exception:
            logger.debug("KB evidence gathering failed for %s", search_term, exc_info=True)

        return evidence

    # ── Classification ───────────────────────────────────────────────────

    async def _classify(self, contact: dict[str, Any]) -> dict[str, Any]:
        raw = await self._delphi.handle(
            "Classify this contact for CRM inclusion.",
            {
                "task": "classify_contact",
                "contact_data": contact,
                "order_history": contact.get("order_history", ""),
                "email_history": contact.get("email_evidence", ""),
            },
        )

        try:
            cleaned = raw.strip()
            if cleaned.startswith("```"):
                lines = cleaned.split("\n")
                lines = [l for l in lines if not l.strip().startswith("```")]
                cleaned = "\n".join(lines)
            return json.loads(cleaned)
        except (json.JSONDecodeError, TypeError):
            logger.warning("Delphi returned non-JSON for %s: %s", contact.get("email"), raw[:200])
            return {"contact_type": "OTHER", "confidence": "LOW", "reasoning": "Parse failure"}

    async def _ensure_company(self, name: str, region: str | None = None) -> str | None:
        companies = await self._crm.list_companies()
        for c in companies:
            if c.name.lower() == name.lower():
                return str(c.id)
        company = await self._crm.create_company(name=name, region=region)
        return str(company.id)

    # ── Data extraction: Gmail ───────────────────────────────────────────

    async def _extract_from_gmail(
        self,
        after: str = "",
        before: str = "",
    ) -> list[dict[str, Any]]:
        """Extract unique contacts from Gmail sent and received messages.

        Paginates through all matching messages (no 500-message cap) and
        extracts richer metadata including subject lines for better
        downstream classification.

        Parameters
        ----------
        after : str
            Start date in ``YYYY/MM/DD`` format (inclusive).
        before : str
            End date in ``YYYY/MM/DD`` format (exclusive).
        """
        try:
            from google.auth.transport.requests import Request
            from google.oauth2.credentials import Credentials
            from googleapiclient.discovery import build

            scopes = ["https://www.googleapis.com/auth/gmail.readonly"]
            token_path = Path("token.json")

            if not token_path.exists():
                logger.warning("No Gmail token.json — skipping Gmail extraction")
                return []

            q_parts: list[str] = []
            if after:
                q_parts.append(f"after:{after}")
            if before:
                q_parts.append(f"before:{before}")
            q = " ".join(q_parts) if q_parts else None

            def _fetch() -> list[dict[str, Any]]:
                creds = Credentials.from_authorized_user_file(str(token_path), scopes)
                if creds and creds.expired and creds.refresh_token:
                    creds.refresh(Request())
                    token_path.write_text(creds.to_json())

                service = build("gmail", "v1", credentials=creds)

                all_stubs: list[dict[str, Any]] = []
                page_token: str | None = None
                while True:
                    kwargs: dict[str, Any] = {"userId": "me", "maxResults": 500}
                    if q:
                        kwargs["q"] = q
                    if page_token:
                        kwargs["pageToken"] = page_token

                    resp = service.users().messages().list(**kwargs).execute()
                    stubs = resp.get("messages", [])
                    all_stubs.extend(stubs)

                    page_token = resp.get("nextPageToken")
                    if not page_token or not stubs:
                        break

                logger.info("Gmail pagination complete: %d message stubs", len(all_stubs))

                contacts_by_email: dict[str, dict[str, Any]] = {}
                email_counts: dict[str, int] = {}
                subjects_by_email: dict[str, list[str]] = {}

                for stub in all_stubs:
                    msg = service.users().messages().get(
                        userId="me", id=stub["id"],
                        format="metadata",
                        metadataHeaders=["From", "To", "Subject"],
                    ).execute()
                    headers = {
                        h["name"].lower(): h["value"]
                        for h in msg.get("payload", {}).get("headers", [])
                    }
                    from_name, from_email = parseaddr(headers.get("from", ""))
                    to_name, to_email = parseaddr(headers.get("to", ""))
                    subject = headers.get("subject", "")

                    for addr_name, addr_email in [(from_name, from_email), (to_name, to_email)]:
                        if not addr_email or "@" not in addr_email:
                            continue
                        addr_email = addr_email.lower().strip()
                        domain = addr_email.split("@", 1)[1]
                        if domain in _OWN_DOMAINS:
                            continue

                        email_counts[addr_email] = email_counts.get(addr_email, 0) + 1

                        if subject:
                            subjects_by_email.setdefault(addr_email, []).append(subject)

                        if addr_email not in contacts_by_email:
                            company_guess = domain.split(".")[0].title()
                            contacts_by_email[addr_email] = {
                                "email": addr_email,
                                "name": addr_name or addr_email.split("@")[0],
                                "company": company_guess,
                                "source": "gmail",
                            }

                _PROPOSAL_RE = _re_module.compile(
                    r"(?i)(quote|proposal|offer|pricing|PF1|PF2|ATF|AM[-\s]|IMG|FCS|"
                    r"thermoform|vacuum\s*form|machine\s+inquiry|techno.?commercial)"
                )

                for email_addr, contact in contacts_by_email.items():
                    count = email_counts.get(email_addr, 0)
                    contact["email_count"] = count
                    contact["has_interactions"] = count > 0

                    subs = subjects_by_email.get(email_addr, [])
                    contact["subjects_sample"] = subs[:5]
                    contact["has_proposal_signal"] = any(
                        _PROPOSAL_RE.search(s) for s in subs
                    )

                return list(contacts_by_email.values())

            results = await asyncio.to_thread(_fetch)
            logger.info("Extracted %d contacts from Gmail", len(results))
            return results

        except Exception:
            logger.exception("Gmail extraction failed")
            return []

    # ── Data extraction: Knowledge Base ──────────────────────────────────

    async def _extract_from_kb(self) -> list[dict[str, Any]]:
        """Extract company/contact mentions from the Qdrant knowledge base."""
        try:
            headers = {"Content-Type": "application/json"}
            if self._qdrant_api_key:
                headers["api-key"] = self._qdrant_api_key

            all_points: list[dict[str, Any]] = []
            offset = None

            async with httpx.AsyncClient(timeout=30) as client:
                while True:
                    body: dict[str, Any] = {
                        "limit": 100,
                        "with_payload": {"include": ["content", "source", "source_category", "metadata"]},
                        "with_vector": False,
                    }
                    if offset is not None:
                        body["offset"] = offset

                    resp = await client.post(
                        f"{self._qdrant_url}/collections/{self._qdrant_collection}/points/scroll",
                        headers=headers,
                        json=body,
                    )
                    resp.raise_for_status()
                    result = resp.json()["result"]
                    points = result["points"]
                    all_points.extend(points)

                    offset = result.get("next_page_offset")
                    if not offset or not points:
                        break

            contacts: dict[str, dict[str, Any]] = {}
            for pt in all_points:
                payload = pt.get("payload", {})
                metadata = payload.get("metadata", {})
                content = payload.get("content", "")
                cat = payload.get("source_category", "")

                customer_name = metadata.get("customer", "")
                if customer_name and isinstance(customer_name, str) and len(customer_name) > 2:
                    key = customer_name.lower().strip()
                    if key not in contacts:
                        has_order = any(kw in cat.lower() for kw in ("order", "po"))
                        has_quote = any(kw in cat.lower() for kw in ("quote", "offer", "proposal"))
                        contacts[key] = {
                            "name": customer_name,
                            "company": customer_name,
                            "email": "",
                            "source": f"kb:{cat}",
                            "has_order": has_order,
                            "has_quote": has_quote,
                            "kb_category": cat,
                        }

            results = [c for c in contacts.values() if c.get("email") or c.get("company")]
            logger.info(
                "Extracted %d contacts from KB (%d with email, %d company-only)",
                len(results),
                sum(1 for c in results if c.get("email")),
                sum(1 for c in results if not c.get("email")),
            )
            return results

        except Exception:
            logger.exception("KB extraction failed")
            return []

    # ── Data extraction: Neo4j ───────────────────────────────────────────

    async def _extract_from_neo4j(self) -> list[dict[str, Any]]:
        try:
            from ira.brain.knowledge_graph import KnowledgeGraph
            graph = KnowledgeGraph()
            query = (
                "MATCH (p:Person) "
                "OPTIONAL MATCH (p)-[:WORKS_AT]->(c:Company) "
                "RETURN p, c LIMIT 500"
            )
            records = await graph.run_cypher(query)

            results: list[dict[str, Any]] = []
            for record in records:
                p_node = record.get("p")
                c_node = record.get("c")
                person = dict(p_node) if p_node and hasattr(p_node, "__iter__") else {}
                company = dict(c_node) if c_node and hasattr(c_node, "__iter__") else {}

                email = person.get("email", "")
                if not email:
                    continue
                results.append({
                    "email": email,
                    "name": person.get("name", ""),
                    "company": company.get("name", ""),
                    "region": company.get("region", ""),
                    "source": "neo4j",
                })
            logger.info("Extracted %d contacts from Neo4j", len(results))
            return results
        except Exception:
            logger.debug("Neo4j extraction failed", exc_info=True)
            return []

    # ── Data extraction: Imports (24_WebSite_Leads) ──────────────────────

    @staticmethod
    def _parse_lead_context_file(path: Path) -> dict[str, Any] | None:
        """Parse a lead*_contact_context.md file; return contact dict or None if no email."""
        text = path.read_text(encoding="utf-8", errors="replace")
        # Extract **Key:** value (value may be on same line or follow)
        key_value: dict[str, str] = {}
        for m in _re_module.finditer(r"\*\*([A-Za-z?]+)\*\*\s*[:\-]?\s*(.+)", text):
            key = m.group(1).strip().lower().replace("?", "")
            val = m.group(2).strip()
            if key and val and key not in key_value:
                key_value[key] = val
        # Also match "- **Key:** value" and value on next line
        lines = text.split("\n")
        for i, line in enumerate(lines):
            if line.strip().startswith("- **") and ":**" in line:
                k, _, v = line.partition(":**")
                key = k.replace("- **", "").strip().lower().replace("?", "")
                val = v.strip()
                if key and key not in key_value:
                    key_value[key] = val or (lines[i + 1].strip() if i + 1 < len(lines) else "")
                elif key and not val and i + 1 < len(lines):
                    key_value[key] = lines[i + 1].strip()

        email = (key_value.get("email") or "").strip()
        if not email or "@" not in email:
            return None

        name = (key_value.get("name") or "").strip() or email.split("@")[0]
        company = (key_value.get("company") or "").strip()
        # Clean company (e.g. "Tricomposite Pty Ltd (Australia)." -> "Tricomposite Pty Ltd")
        if company and company.endswith("."):
            company = company[:-1].strip()
        region_raw = (key_value.get("region") or "").strip()
        # Normalize region to first part (e.g. "Asia-Pacific (Australia)" -> "Asia-Pacific")
        region = region_raw.split("(")[0].strip() if region_raw else None
        if region and len(region) > 80:
            region = region[:80]
        machine = (key_value.get("machine to offer") or key_value.get("machine") or "").strip()
        if not machine and "machine" in key_value:
            machine = key_value.get("machine", "").strip()
        # First machine model pattern (e.g. PF1-C-3030, PF1-X-6520)
        machine_match = _re_module.search(r"PF1[-\s]?[CX]?[-\s]?\d{4,}", machine or "")
        machine_model = machine_match.group(0).replace(" ", "-") if machine_match else (machine[:100] if machine else None)
        inquiry = (key_value.get("inquiry") or key_value.get("inquiry (form)") or "").strip()
        segment = (key_value.get("segment") or "").strip()
        application = segment or (inquiry[:200] if inquiry else None)

        client_line = (key_value.get("client") or "").lower()
        body_lower = text.lower()
        if "existing customer" in body_lower or "they placed" in body_lower or "we delivered" in body_lower or "client? yes" in body_lower:
            contact_type_hint = "LIVE_CUSTOMER"
        elif "past customer" in body_lower or "past contact" in body_lower and "client? no" in body_lower:
            contact_type_hint = "PAST_CUSTOMER"
        elif "client? no" in body_lower and ("we sent" in body_lower or "quote" in body_lower or "offer" in body_lower):
            contact_type_hint = "LEAD_WITH_INTERACTIONS"
        elif "client? no" in body_lower or "lead" in body_lower:
            contact_type_hint = "LEAD_NO_INTERACTIONS"
        else:
            contact_type_hint = "LEAD_WITH_INTERACTIONS"

        return {
            "email": email,
            "name": name,
            "company": company or name,
            "region": region,
            "machine_model": machine_model,
            "application": application,
            "contact_type_hint": contact_type_hint,
            "source": "imports",
        }

    async def _extract_from_imports(self) -> list[dict[str, Any]]:
        """Extract contacts from data/imports/24_WebSite_Leads lead*_contact_context.md files."""
        contacts: list[dict[str, Any]] = []
        if not _IMPORTS_LEADS_DIR.is_dir():
            logger.warning("Imports leads dir not found: %s", _IMPORTS_LEADS_DIR)
            return contacts
        for path in sorted(_IMPORTS_LEADS_DIR.glob("*_contact_context.md")):
            try:
                parsed = self._parse_lead_context_file(path)
                if parsed:
                    contacts.append(parsed)
            except Exception:
                logger.debug("Failed to parse %s", path.name, exc_info=True)
        logger.info("Extracted %d contacts from imports (24_WebSite_Leads)", len(contacts))
        return contacts

    # ── Data extraction: Imports (07_Leads_and_Contacts) ─────────────────

    @staticmethod
    def _normalize_header(h: str) -> str:
        return (h or "").strip().lower().replace(" ", "_").replace("-", "_")

    @staticmethod
    def _map_row_to_contact(row: dict[str, Any], headers_map: dict[str, str], source: str) -> dict[str, Any] | None:
        """Map a row dict (keyed by normalized header) to contact dict. Requires email."""
        email = (
            (row.get("email") or row.get("e_mail") or row.get("contact_email") or "")
            .strip()
        )
        if isinstance(email, (int, float)):
            email = str(int(email)) if isinstance(email, float) and email == int(email) else str(email)
        email = email.strip().lower()
        if not email or "@" not in email:
            return None
        name = (row.get("name") or row.get("contact") or row.get("contact_name") or row.get("full_name") or "").strip()
        if isinstance(name, (int, float)):
            name = str(name)
        if not name:
            name = email.split("@")[0]
        company = (row.get("company") or row.get("organisation") or row.get("organization") or "").strip()
        if isinstance(company, (int, float)):
            company = str(company)
        if not company:
            company = name
        region = (row.get("region") or row.get("country") or row.get("location") or "").strip()
        if isinstance(region, (int, float)):
            region = str(region)
        if region and len(region) > 100:
            region = region[:100]
        return {
            "email": email,
            "name": name,
            "company": company,
            "region": region or None,
            "source": source,
        }

    def _read_xlsx_contacts(self, path: Path, source_label: str) -> list[dict[str, Any]]:
        """Read first sheet of an XLSX; first row = headers. Return list of contact dicts."""
        contacts: list[dict[str, Any]] = []
        try:
            from openpyxl import load_workbook
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            if not ws:
                return contacts
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) < 2:
                return contacts
            raw_headers = [str(c).strip() if c is not None else "" for c in rows[0]]
            col_to_norm: dict[int, str] = {}
            email_cols = ("email", "e_mail", "contact_email", "e-mail")
            for i, h in enumerate(raw_headers):
                n = self._normalize_header(h)
                if n in email_cols:
                    col_to_norm[i] = "email"
                elif n in ("name", "contact", "contact_name", "full_name"):
                    col_to_norm[i] = "name"
                elif n in ("company", "organisation", "organization", "company_name"):
                    col_to_norm[i] = "company"
                elif n in ("region", "country", "location"):
                    col_to_norm[i] = "region"
            if "email" not in col_to_norm.values():
                return contacts
            for row in rows[1:]:
                if not row:
                    continue
                rdict: dict[str, Any] = {}
                for i, val in enumerate(row):
                    if i in col_to_norm and val is not None and str(val).strip():
                        rdict[col_to_norm[i]] = str(val).strip()
                if "email" not in rdict or "@" not in str(rdict.get("email", "")):
                    continue
                contact = self._map_row_to_contact(rdict, {}, source_label)
                if contact:
                    contacts.append(contact)
            wb.close()
        except Exception:
            logger.debug("Failed to read XLSX %s", path.name, exc_info=True)
        return contacts

    def _read_csv_contacts(self, path: Path, source_label: str) -> list[dict[str, Any]]:
        """Read CSV; first row = headers. Return list of contact dicts."""
        contacts: list[dict[str, Any]] = []
        try:
            with path.open(encoding="utf-8", errors="replace") as f:
                reader = csv.DictReader(f)
                if not reader.fieldnames:
                    return contacts
                for row in reader:
                    norm_row: dict[str, Any] = {}
                    for k, v in row.items():
                        n = self._normalize_header(k)
                        if v is None or (isinstance(v, str) and not v.strip()):
                            continue
                        if "email" in n or n == "e_mail":
                            norm_row["email"] = v.strip()
                        elif "name" in n or n in ("contact", "contact_name", "full_name"):
                            norm_row["name"] = v.strip()
                        elif "company" in n or "organisation" in n:
                            norm_row["company"] = v.strip()
                        elif n in ("region", "country", "location"):
                            norm_row["region"] = v.strip()
                    if norm_row.get("email") and "@" in str(norm_row["email"]):
                        contact = self._map_row_to_contact(norm_row, {}, source_label)
                        if contact:
                            contacts.append(contact)
        except Exception:
            logger.debug("Failed to read CSV %s", path.name, exc_info=True)
        return contacts

    async def _extract_from_imports_07(self) -> list[dict[str, Any]]:
        """Extract contacts from data/imports/07_Leads_and_Contacts (XLSX/CSV) and Single Station Inquiry at imports root."""
        contacts: list[dict[str, Any]] = []
        # 07_Leads_and_Contacts
        if _IMPORTS_07_DIR.is_dir():
            for path in sorted(_IMPORTS_07_DIR.glob("*.xlsx")) + sorted(_IMPORTS_07_DIR.glob("*.xls")):
                contacts.extend(self._read_xlsx_contacts(path, "imports_07"))
            for path in sorted(_IMPORTS_07_DIR.glob("*.csv")):
                contacts.extend(self._read_csv_contacts(path, "imports_07"))
        # Single Station Inquiry at imports root
        single_xlsx = _IMPORTS_ROOT / "Single Station Inquiry Form (Responses).xlsx"
        if single_xlsx.exists():
            contacts.extend(self._read_xlsx_contacts(single_xlsx, "imports_07"))
        logger.info("Extracted %d contacts from imports_07 (07_Leads + Single Station)", len(contacts))
        return contacts

    # ── Data extraction: Takeout sent mail ───────────────────────────────

    async def _extract_from_takeout_sent(self) -> list[dict[str, Any]]:
        """Extract contacts we've emailed from data/takeout_ingest/*.mbox (From: *@machinecraft* -> To:)."""
        contacts: list[dict[str, Any]] = []
        if not _TAKEOUT_INGEST_DIR.is_dir():
            logger.warning("Takeout ingest dir not found: %s", _TAKEOUT_INGEST_DIR)
            return contacts
        seen_emails: set[str] = set()
        mbox_paths = list(_TAKEOUT_INGEST_DIR.glob("*.mbox"))
        for mbox_path in mbox_paths[:30]:
            try:
                mbox = mailbox.mbox(str(mbox_path))
                for msg in mbox:
                    from_raw = str(msg.get("From", ""))
                    from_name, from_email = parseaddr(from_raw)
                    from_email = (from_email or "").strip().lower()
                    if not from_email or "@" not in from_email:
                        continue
                    domain = from_email.split("@", 1)[1]
                    if domain not in _OWN_DOMAINS:
                        continue
                    to_raw = str(msg.get("To", ""))
                    for _name, addr in getaddresses([to_raw]):
                        email = (addr or "").strip().lower()
                        if not email or "@" not in email or email in seen_emails:
                            continue
                        if email.split("@", 1)[1] in _OWN_DOMAINS:
                            continue
                        if _is_obviously_not_business(email):
                            continue
                        seen_emails.add(email)
                        name = (_name or "").strip() or email.split("@")[0]
                        company_guess = email.split("@", 1)[1].split(".")[0].title()
                        contacts.append({
                            "email": email,
                            "name": name,
                            "company": company_guess,
                            "source": "takeout_sent",
                        })
                        if len(contacts) >= _MAX_TAKEOUT_SENT_CONTACTS:
                            break
                mbox.close()
                if len(contacts) >= _MAX_TAKEOUT_SENT_CONTACTS:
                    break
            except Exception:
                logger.debug("Failed to read mbox %s", mbox_path.name, exc_info=True)
        logger.info("Extracted %d contacts from takeout sent mail", len(contacts))
        return contacts

    # ── Deduplication ────────────────────────────────────────────────────

    def _deduplicate(self, contacts: list[dict[str, Any]]) -> list[dict[str, Any]]:
        by_email: dict[str, dict[str, Any]] = {}
        for c in contacts:
            email = (c.get("email") or "").strip().lower()
            if not email:
                continue
            if email not in by_email:
                by_email[email] = c
            else:
                existing = by_email[email]
                for key in ("company", "name", "role", "region", "phone"):
                    if not existing.get(key) and c.get(key):
                        existing[key] = c[key]
                for key in ("machine_model", "application", "contact_type_hint", "source"):
                    if c.get(key) and not existing.get(key):
                        existing[key] = c[key]
                for key in ("has_interactions", "has_order", "has_quote"):
                    if c.get(key):
                        existing[key] = True
                if c.get("email_count", 0) > existing.get("email_count", 0):
                    existing["email_count"] = c["email_count"]
        return list(by_email.values())
