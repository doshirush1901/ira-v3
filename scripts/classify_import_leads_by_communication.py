#!/usr/bin/env python3
"""Classify leads from import files by whether they have ever communicated with us.

Communication = they sent us at least one email (reply or new) that is not an auto-reply.
We do NOT count "we sent, they didn't reply" as communicated.

Sources (all under data/imports):
- 06_Market_Research_and_Analysis/Competitor Customer Analysis - Single Station Thermoforming.xlsx
- 06_Market_Research_and_Analysis/LLM_Thermoforming_Prospects.xlsx
- 06_Market_Research_and_Analysis/Thermoforming Machine Market Data - Europe.xlsx
- 07_Leads_and_Contacts/European & US Contacts for Single Station Nov 203.csv
- 07_Leads_and_Contacts/K2016 - Data for Exhibition Inquiry & Analysis.xlsx
- 07_Leads_and_Contacts/K2022 Meetings.xlsx
- 07_Leads_and_Contacts/K2025 Leads.xlsx
- 02_Orders_and_POs/Top 50 European Thermoforming Companies – Profiles & Opportunities.pdf (skipped; no email extraction)

Usage:
  # With Ira API running (for Gmail search):
  poetry run python scripts/classify_import_leads_by_communication.py
  poetry run python scripts/classify_import_leads_by_communication.py --output data/reports/leads_communicated_only.csv
  poetry run python scripts/classify_import_leads_by_communication.py --dry-run   # only extract leads, no API calls

  # Output: prints communicated-only table; with --output writes CSV (and optionally full CSV with communicated=True/False).
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

# Optional: httpx for API, openpyxl for XLSX
try:
    import httpx
except ImportError:
    httpx = None
try:
    import openpyxl
except ImportError:
    openpyxl = None

AUTO_REPLY_PATTERNS = [
    r"out of (the )?office",
    r"automatic reply",
    r"auto[- ]?reply",
    r"auto[- ]?response",
    r"automatische antwort",
    r"réponse automatique",
    r"abwesenheitsnotiz",
    r"vacation",
    r"holiday",
    r"away from (my )?desk",
    r"maternity leave",
    r"paternity leave",
    r"currently unavailable",
    r"i am currently out",
    r"thank you for your (email|message).*will (get back|respond|reply)",
    r"this is an automated",
    r"do not reply",
    r"noreply",
    r"no-reply",
    r"mailer-daemon",
    r"postmaster",
    r"delivery (status|failure|failed)",
    r"undeliverable",
    r"returned mail",
    r"bounce",
]
AUTO_REPLY_REGEX = re.compile("|".join(AUTO_REPLY_PATTERNS), re.IGNORECASE)
EMAIL_REGEX = re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+")

IMPORTS = PROJECT_ROOT / "data" / "imports"
API_URL = "http://localhost:8000"
API_KEY = ""  # set from env in main


def _load_env() -> None:
    import os
    from pathlib import Path
    env = PROJECT_ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#"):
                continue
            if "=" in line:
                k, v = line.split("=", 1)
                k, v = k.strip(), v.strip().strip('"').strip("'")
                if k == "API_SECRET_KEY":
                    globals()["API_KEY"] = v
                elif k == "IRA_API_URL":
                    globals()["API_URL"] = v.strip()


def _normalize_email(s: str | None) -> str | None:
    if not s or "@" not in s:
        return None
    s = s.strip().lower()
    if not EMAIL_REGEX.match(s):
        return None
    return s


def _is_auto_reply(subject: str, body: str, from_addr: str) -> bool:
    subject = (subject or "").lower()
    body = (body or "")[:500].lower()
    from_addr = (from_addr or "").lower()
    if AUTO_REPLY_REGEX.search(subject):
        return True
    if AUTO_REPLY_REGEX.search(body):
        return True
    if any(x in from_addr for x in ["noreply", "no-reply", "mailer-daemon", "postmaster"]):
        return True
    return False


# ---------------------------------------------------------------------------
# Lead extraction per source
# ---------------------------------------------------------------------------


def _extract_email_from_row(row: list[Any], email_col_index: int) -> str | None:
    if email_col_index < 0 or email_col_index >= len(row):
        return None
    v = row[email_col_index]
    if v is None:
        return None
    return _normalize_email(str(v).strip())


def _extract_emails_from_any_cells(row: list[Any]) -> list[str]:
    out: list[str] = []
    for v in row:
        if v is None:
            continue
        s = str(v).strip()
        for m in EMAIL_REGEX.finditer(s):
            e = _normalize_email(m.group(0))
            if e:
                out.append(e)
    return out


def load_european_us_csv() -> list[dict[str, str]]:
    path = IMPORTS / "07_Leads_and_Contacts" / "European & US Contacts for Single Station Nov 203.csv"
    if not path.exists():
        return []
    rows: list[dict[str, str]] = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        r = csv.DictReader(f)
        for row in r:
            email = _normalize_email(row.get("Email Address") or row.get("Email") or "")
            if not email:
                continue
            rows.append({
                "email": email,
                "name": f"{row.get('First Name', '').strip()} {row.get('Last Name', '').strip()}".strip() or "—",
                "company": (row.get("Company Name") or "").strip() or "—",
                "source": path.name,
            })
    return rows


def load_xlsx_simple(
    path: Path,
    email_col: str | int,
    company_col: str | int | None = None,
    name_col: str | int | None = None,
    header_row: int = 1,
) -> list[dict[str, str]]:
    """Load XLSX; email_col can be 1-based column name (from row 1) or 0-based index."""
    if not openpyxl or not path.exists():
        return []
    rows: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    if not sh:
        wb.close()
        return []
    all_rows = list(sh.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    headers = [str(h).strip() if h is not None else "" for h in all_rows[header_row - 1]]
    # resolve column index
    if isinstance(email_col, str):
        try:
            email_idx = headers.index(email_col)
        except ValueError:
            email_idx = next((i for i, h in enumerate(headers) if h and "email" in h.lower()), -1)
    else:
        email_idx = email_col
    company_idx: int = -1
    if company_col is not None:
        if isinstance(company_col, str):
            try:
                company_idx = headers.index(company_col)
            except ValueError:
                company_idx = -1
        else:
            company_idx = company_col
    name_idx: int = -1
    if name_col is not None:
        if isinstance(name_col, str):
            try:
                name_idx = headers.index(name_col)
            except ValueError:
                name_idx = -1
        else:
            name_idx = name_col

    for r in all_rows[header_row:]:
        row = list(r) if r else []
        email = _extract_email_from_row(row, email_idx) if email_idx >= 0 else None
        if not email:
            # try any cell containing @
            emails = _extract_emails_from_any_cells(row)
            email = emails[0] if emails else None
        if not email:
            continue
        company = (row[company_idx] if 0 <= company_idx < len(row) and row[company_idx] else None) or "—"
        name = (row[name_idx] if 0 <= name_idx < len(row) and row[name_idx] else None) or "—"
        if isinstance(company, str):
            company = company.strip() or "—"
        else:
            company = str(company).strip() if company else "—"
        if isinstance(name, str):
            name = name.strip() or "—"
        else:
            name = str(name).strip() if name else "—"
        rows.append({
            "email": email,
            "name": name or "—",
            "company": company or "—",
            "source": path.name,
        })
    return rows


def load_k2025() -> list[dict[str, str]]:
    path = IMPORTS / "07_Leads_and_Contacts" / "K2025 Leads.xlsx"
    # headers: None, Company, Email, Country, Person, ...
    return load_xlsx_simple(path, email_col=2, company_col=1, name_col=4, header_row=1)


def load_k2022() -> list[dict[str, str]]:
    path = IMPORTS / "07_Leads_and_Contacts" / "K2022 Meetings.xlsx"
    # Mail Sent, Quote, None, Client, Email, First Name, Last Name, ...
    # We use company (Client) for both; name from First+Last is done in generic loader if we had two name cols
    return load_xlsx_simple(path, email_col=4, company_col=3, name_col=5, header_row=1)


def load_k2016() -> list[dict[str, str]]:
    path = IMPORTS / "07_Leads_and_Contacts" / "K2016 - Data for Exhibition Inquiry & Analysis.xlsx"
    # None, Customer Name, Region, Email, ...
    return load_xlsx_simple(path, email_col=3, company_col=1, name_col=1, header_row=1)


def load_llm_prospects() -> list[dict[str, str]]:
    path = IMPORTS / "06_Market_Research_and_Analysis" / "LLM_Thermoforming_Prospects.xlsx"
    return load_xlsx_simple(path, email_col="Primary_Email", company_col="Company_Name", name_col=None, header_row=1)


def load_competitor_analysis() -> list[dict[str, str]]:
    """Competitor Customer Analysis has 'Contact / website' - no clear email column; scan all cells."""
    path = IMPORTS / "06_Market_Research_and_Analysis" / "Competitor Customer Analysis - Single Station Thermoforming.xlsx"
    if not openpyxl or not path.exists():
        return []
    rows: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    if not sh:
        wb.close()
        return []
    all_rows = list(sh.iter_rows(values_only=True))
    wb.close()
    # Row 0 = headers: Country, Address, Contact / website, Company Est., ...
    headers = [str(h).strip() if h is not None else "" for h in (all_rows[0] or [])]
    country_idx = next((i for i, h in enumerate(headers) if "country" in h.lower()), 0)
    contact_idx = next((i for i, h in enumerate(headers) if "contact" in h.lower() or "website" in h.lower()), 2)
    company_idx = next((i for i, h in enumerate(headers) if "company" in h.lower()), 1)
    for r in (all_rows[1:] or []):
        row = list(r) if r else []
        emails = _extract_emails_from_any_cells(row)
        for email in emails[:1]:  # first email per row
            company = (row[company_idx] if 0 <= company_idx < len(row) else None) or "—"
            country = (row[country_idx] if 0 <= country_idx < len(row) else None) or ""
            rows.append({
                "email": email,
                "name": "—",
                "company": str(company).strip() if company else "—",
                "source": path.name,
            })
            break
    return rows


def load_europe_market_data() -> list[dict[str, str]]:
    """Thermoforming Machine Market Data - Europe: headers in row 2 (0-based), data from row 3."""
    path = IMPORTS / "06_Market_Research_and_Analysis" / "Thermoforming Machine Market Data - Europe.xlsx"
    if not openpyxl or not path.exists():
        return []
    rows: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    if not sh:
        wb.close()
        return []
    all_rows = list(sh.iter_rows(values_only=True))
    wb.close()
    # Find header row (first row containing 'Email')
    header_row = 0
    for i, row in enumerate(all_rows):
        if not row:
            continue
        if any("email" in str(c).lower() for c in row if c):
            header_row = i
            break
    headers = [str(h).strip() if h is not None else "" for h in (all_rows[header_row] or [])]
    email_idx = next((i for i, h in enumerate(headers) if "email" in h.lower()), -1)
    name_idx = next((i for i, h in enumerate(headers) if "person" in h.lower() or "company" in h.lower()), 0)
    country_idx = next((i for i, h in enumerate(headers) if "country" in h.lower()), -1)
    for r in all_rows[header_row + 1:]:
        row = list(r) if r else []
        email = _extract_email_from_row(row, email_idx) if email_idx >= 0 else None
        if not email:
            emails = _extract_emails_from_any_cells(row)
            email = emails[0] if emails else None
        if not email:
            continue
        name = (row[name_idx] if 0 <= name_idx < len(row) else None) or "—"
        rows.append({
            "email": email,
            "name": str(name).strip() if name else "—",
            "company": str(name).strip() if name else "—",
            "source": path.name,
        })
    return rows


def load_all_leads() -> list[dict[str, str]]:
    """Load from all configured sources and dedupe by email (keep first occurrence)."""
    seen: set[str] = set()
    out: list[dict[str, str]] = []
    loaders = [
        ("European & US Contacts", load_european_us_csv),
        ("K2025 Leads", load_k2025),
        ("K2022 Meetings", load_k2022),
        ("K2016", load_k2016),
        ("LLM_Thermoforming_Prospects", load_llm_prospects),
        ("Competitor Customer Analysis", load_competitor_analysis),
        ("Thermoforming Machine Market Data - Europe", load_europe_market_data),
    ]
    for _label, fn in loaders:
        try:
            for lead in fn():
                e = lead["email"]
                if e not in seen:
                    seen.add(e)
                    out.append(lead)
        except Exception as e:
            print(f"[warn] {_label}: {e}", file=sys.stderr)
    return out


# ---------------------------------------------------------------------------
# Gmail check via Ira API
# ---------------------------------------------------------------------------


def search_emails_from(from_address: str, max_results: int = 50) -> list[dict[str, Any]]:
    """Search Gmail for emails we received FROM this address (they replied or sent new)."""
    if not httpx:
        return []
    headers = {"Content-Type": "application/json"}
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"
    payload = {
        "from_address": from_address.strip(),
        "to_address": "",
        "max_results": max_results,
    }
    try:
        r = httpx.post(f"{API_URL}/api/email/search", json=payload, headers=headers, timeout=30)
        r.raise_for_status()
        data = r.json()
        return data.get("emails", [])
    except Exception as e:
        print(f"  [api error] {from_address}: {e}", file=sys.stderr)
        return []


def has_communicated(email: str) -> tuple[bool, int, int]:
    """Returns (communicated, total_from_them, genuine_count)."""
    emails = search_emails_from(email)
    if not emails:
        return False, 0, 0
    genuine = 0
    for e in emails:
        subj = e.get("subject", "")
        body = (e.get("body") or "")[:500]
        from_addr = e.get("from", "")
        if _is_auto_reply(subj, body, from_addr):
            continue
        genuine += 1
    return genuine > 0, len(emails), genuine


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def main() -> None:
    import argparse
    _load_env()
    parser = argparse.ArgumentParser(description="Classify import leads by communication (they replied or sent email).")
    parser.add_argument("--output", "-o", type=Path, help="Write communicated-only CSV here.")
    parser.add_argument("--full", type=Path, help="Write full classification CSV (all leads + communicated column).")
    parser.add_argument("--dry-run", action="store_true", help="Only extract leads; do not call API.")
    parser.add_argument("--limit", type=int, default=0, help="Limit number of leads to check (0 = all).")
    args = parser.parse_args()

    print("Loading leads from import files...", flush=True)
    leads = load_all_leads()
    print(f"  Total unique leads (by email): {len(leads)}", flush=True)
    if not leads:
        print("No leads found. Check paths under data/imports.", file=sys.stderr)
        return

    if args.dry_run:
        print("\n[DRY RUN] First 20 leads (email, company, source):")
        for L in leads[:20]:
            print(f"  {L['email']} | {L['company']} | {L['source']}")
        if len(leads) > 20:
            print(f"  ... and {len(leads) - 20} more")
        return

    if not httpx:
        print("Install httpx to call Ira API: poetry add httpx", file=sys.stderr)
        return

    to_check = leads if args.limit <= 0 else leads[: args.limit]
    communicated_only: list[dict[str, str | int]] = []
    full_rows: list[dict[str, Any]] = []

    for i, lead in enumerate(to_check, 1):
        email = lead["email"]
        print(f"[{i}/{len(to_check)}] {email} ({lead['company']})...", flush=True)
        ok, total, genuine = has_communicated(email)
        row = {**lead, "communicated": ok, "emails_from_them": total, "genuine_replies": genuine}
        full_rows.append(row)
        if ok:
            communicated_only.append(row)
            print(f"  -> communicated (genuine: {genuine})")
        else:
            print(f"  -> no communication" + (f" (from_them: {total})" if total else ""))

    print(f"\n--- Summary ---")
    print(f"Checked: {len(to_check)}")
    print(f"Communicated (replied or sent us email): {len(communicated_only)}")

    if communicated_only:
        print("\n--- Communicated only (they replied or sent us email) ---")
        for r in communicated_only:
            print(f"  {r['email']} | {r['company']} | {r['source']} | genuine={r.get('genuine_replies', 0)}")

    if args.output and communicated_only:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["email", "name", "company", "source", "emails_from_them", "genuine_replies"])
            w.writeheader()
            w.writerows({k: r.get(k, "") for k in ["email", "name", "company", "source", "emails_from_them", "genuine_replies"]} for r in communicated_only)
        print(f"\nWrote communicated-only CSV: {args.output}")

    if args.full and full_rows:
        args.full.parent.mkdir(parents=True, exist_ok=True)
        with open(args.full, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=["email", "name", "company", "source", "communicated", "emails_from_them", "genuine_replies"])
            w.writeheader()
            w.writerows(full_rows)
        print(f"Wrote full classification CSV: {args.full}")


if __name__ == "__main__":
    main()
