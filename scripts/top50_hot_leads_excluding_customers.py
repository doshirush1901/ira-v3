#!/usr/bin/env python3
"""Top 50 hottest leads from 07_Leads_and_Contacts.

1. Load all customers (past/present/new) from:
   - List of Customers - Machinecraft.xlsx
   - Clients MC EUROPE.xlsx
2. Load all leads from 07_Leads_and_Contacts (all XLSX/CSV except the two customer files).
3. Exclude any lead whose email or company is in the customer lists.
4. Keep only leads who have communicated: they emailed us (or we emailed them) and we heard back
   (they replied or asked for quote / gave feedback). Uses Ira API email search.
5. Rank by hottest (genuine reply count, then total emails from them); output top 50.

Usage:
  # With Ira API running (for Gmail check):
  poetry run python scripts/top50_hot_leads_excluding_customers.py
  poetry run python scripts/top50_hot_leads_excluding_customers.py --output data/reports/top50_hot_leads.csv
  poetry run python scripts/top50_hot_leads_excluding_customers.py --dry-run   # exclude customers only; no API
  poetry run python scripts/top50_hot_leads_excluding_customers.py --limit 100 # check first 100 leads only
"""

from __future__ import annotations

import csv
import re
import sys
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(PROJECT_ROOT))

IMPORTS_07 = PROJECT_ROOT / "data" / "imports" / "07_Leads_and_Contacts"
API_URL = "http://localhost:8000"
API_KEY = ""

try:
    import httpx
except ImportError:
    httpx = None
try:
    import openpyxl
except ImportError:
    openpyxl = None

EMAIL_REGEX = re.compile(r"[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+")
CUSTOMER_FILES = {
    "List of Customers - Machinecraft.xlsx",
    "Clients MC EUROPE.xlsx",
}


def _normalize_email(s: str | None) -> str | None:
    if not s or "@" not in s:
        return None
    s = s.strip().lower()
    if not EMAIL_REGEX.match(s):
        return None
    return s


def _normalize_company(s: str | None) -> str:
    if not s:
        return ""
    return re.sub(r"\s+", " ", str(s).strip().lower())


# ---------------------------------------------------------------------------
# Customer loading (exclusion set)
# ---------------------------------------------------------------------------


def load_customers_list_of_customers() -> tuple[set[str], set[str]]:
    """List of Customers - Machinecraft.xlsx: headers Company, Email, Person, Region."""
    emails: set[str] = set()
    companies: set[str] = set()
    path = IMPORTS_07 / "List of Customers - Machinecraft.xlsx"
    if not openpyxl or not path.exists():
        return emails, companies
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        return emails, companies
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if len(rows) < 2:
        return emails, companies
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_company = next((i for i, h in enumerate(headers) if "company" in h.lower()), -1)
    col_email = next((i for i, h in enumerate(headers) if "email" in h.lower()), -1)
    for row in rows[1:]:
        if not row:
            continue
        row = list(row)
        if col_email >= 0 and col_email < len(row) and row[col_email]:
            e = _normalize_email(str(row[col_email]))
            if e:
                emails.add(e)
        if col_company >= 0 and col_company < len(row) and row[col_company]:
            c = _normalize_company(row[col_company])
            if c and len(c) > 1:
                companies.add(c)
    return emails, companies


def load_customers_clients_europe() -> tuple[set[str], set[str]]:
    """Clients MC EUROPE.xlsx: no header; col0=name, col1=country, col2=email, col3=website."""
    emails: set[str] = set()
    companies: set[str] = set()
    path = IMPORTS_07 / "Clients MC EUROPE.xlsx"
    if not openpyxl or not path.exists():
        return emails, companies
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    if not ws:
        wb.close()
        return emails, companies
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    for row in rows:
        if not row or len(row) < 3:
            continue
        row = list(row)
        e = _normalize_email(str(row[2]).strip() if row[2] else None)
        if e:
            emails.add(e)
        # Use name or domain as company hint for exclusion
        name = str(row[0]).strip() if len(row) > 0 and row[0] else ""
        if name and len(name) > 1 and not name.replace(".", "").isdigit():
            companies.add(_normalize_company(name))
    return emails, companies


def load_all_customers() -> tuple[set[str], set[str]]:
    e1, c1 = load_customers_list_of_customers()
    e2, c2 = load_customers_clients_europe()
    return e1 | e2, c1 | c2


# ---------------------------------------------------------------------------
# Lead loading from 07_Leads (all XLSX/CSV except customer files)
# ---------------------------------------------------------------------------


def _extract_emails_from_row(row: list[Any]) -> list[str]:
    out: list[str] = []
    for v in row:
        if v is None:
            continue
        for m in EMAIL_REGEX.finditer(str(v)):
            e = _normalize_email(m.group(0))
            if e:
                out.append(e)
    return out


def load_xlsx_auto(path: Path) -> list[dict[str, str]]:
    """Load any XLSX: find header row (first row with 'email'), then Email and Company columns."""
    if not openpyxl or not path.exists():
        return []
    out: list[dict[str, str]] = []
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sh = wb.active
    if not sh:
        wb.close()
        return []
    all_rows = list(sh.iter_rows(values_only=True))
    wb.close()
    if not all_rows:
        return []
    header_row_idx = 0
    for i, row in enumerate(all_rows):
        if not row:
            continue
        headers = [str(h).strip().lower() if h is not None else "" for h in row]
        if any(h and "email" in h for h in headers):
            header_row_idx = i
            break
    headers = [str(h).strip().lower() if h is not None else "" for h in all_rows[header_row_idx]]
    email_idx = next((i for i, h in enumerate(headers) if h and "email" in h), -1)
    company_idx = next(
        (i for i, h in enumerate(headers) if h and ("company" in h or "client" in h or "customer" in h or "organisation" in h or "organization" in h)),
        -1,
    )
    name_idx = next(
        (i for i, h in enumerate(headers) if h and ("name" in h or "person" in h or "contact" in h) and "company" not in h),
        -1,
    )
    for r in all_rows[header_row_idx + 1 :]:
        row = list(r) if r else []
        if email_idx >= 0 and email_idx < len(row) and row[email_idx]:
            email = _normalize_email(str(row[email_idx]))
        else:
            emails = _extract_emails_from_row(row)
            email = emails[0] if emails else None
        if not email:
            continue
        company = "—"
        if 0 <= company_idx < len(row) and row[company_idx]:
            company = (str(row[company_idx]).strip() or "—")[:200]
        name = "—"
        if 0 <= name_idx < len(row) and row[name_idx]:
            name = (str(row[name_idx]).strip() or "—")[:200]
        out.append({"email": email, "name": name, "company": company, "source": path.name})
    return out


def load_csv_auto(path: Path) -> list[dict[str, str]]:
    """Load CSV; first row headers; find email and company columns."""
    if not path.exists():
        return []
    rows: list[dict[str, str]] = []
    with open(path, newline="", encoding="utf-8", errors="replace") as f:
        reader = csv.DictReader(f)
        if not reader.fieldnames:
            return rows
        fieldnames = [k.strip().lower() for k in reader.fieldnames]
        email_key = next((k for k in reader.fieldnames if "email" in k.strip().lower()), None)
        company_key = next(
            (k for k in reader.fieldnames if any(x in k.strip().lower() for x in ("company", "client", "organisation", "organization"))),
            None,
        )
        name_key = next(
            (k for k in reader.fieldnames if "name" in k.strip().lower() and "company" not in k.strip().lower()),
            None,
        )
        for row in reader:
            email = _normalize_email(row.get(email_key or "", ""))
            if not email:
                for k, v in row.items():
                    if v and "@" in str(v):
                        email = _normalize_email(str(v))
                        break
            if not email:
                continue
            company = (row.get(company_key or "", "") or "—").strip()[:200] if company_key else "—"
            name = (row.get(name_key or "", "") or "—").strip()[:200] if name_key else "—"
            rows.append({"email": email, "name": name, "company": company or "—", "source": path.name})
    return rows


def load_all_leads_07_only() -> list[dict[str, str]]:
    """All leads from 07_Leads_and_Contacts, excluding customer-list files and PDFs."""
    seen: set[str] = set()
    out: list[dict[str, str]] = []
    if not IMPORTS_07.is_dir():
        return out
    for path in sorted(IMPORTS_07.glob("*.xlsx")) + sorted(IMPORTS_07.glob("*.xls")):
        if path.name in CUSTOMER_FILES:
            continue
        for lead in load_xlsx_auto(path):
            if lead["email"] not in seen:
                seen.add(lead["email"])
                out.append(lead)
    for path in sorted(IMPORTS_07.glob("*.csv")):
        for lead in load_csv_auto(path):
            if lead["email"] not in seen:
                seen.add(lead["email"])
                out.append(lead)
    return out


# ---------------------------------------------------------------------------
# Communication check (Ira API)
# ---------------------------------------------------------------------------

AUTO_REPLY_PATTERNS = [
    r"out of (the )?office", r"automatic reply", r"auto[- ]?reply", r"vacation",
    r"holiday", r"away from (my )?desk", r"do not reply", r"noreply", r"mailer-daemon",
    r"postmaster", r"delivery (status|failure)", r"undeliverable", r"bounce",
]
AUTO_REPLY_REGEX = re.compile("|".join(AUTO_REPLY_PATTERNS), re.IGNORECASE)


def _is_auto_reply(subject: str, body: str, from_addr: str) -> bool:
    if AUTO_REPLY_REGEX.search((subject or "").lower()):
        return True
    if AUTO_REPLY_REGEX.search((body or "")[:500].lower()):
        return True
    if any(x in (from_addr or "").lower() for x in ["noreply", "no-reply", "mailer-daemon", "postmaster"]):
        return True
    return False


def search_emails_from(from_address: str, max_results: int = 50) -> list[dict[str, Any]]:
    if not httpx:
        return []
    headers = {"Content-Type": "application/json"}
    if API_KEY:
        headers["Authorization"] = f"Bearer {API_KEY}"
    try:
        r = httpx.post(
            f"{API_URL}/api/email/search",
            json={"from_address": from_address.strip(), "max_results": max_results},
            headers=headers,
            timeout=30,
        )
        r.raise_for_status()
        return r.json().get("emails", [])
    except Exception:
        return []


def has_communicated(email: str) -> tuple[bool, int, int]:
    """(communicated, total_from_them, genuine_count)."""
    emails = search_emails_from(email)
    if not emails:
        return False, 0, 0
    genuine = 0
    for e in emails:
        if _is_auto_reply(e.get("subject", ""), (e.get("body") or "")[:500], e.get("from", "")):
            continue
        genuine += 1
    return genuine > 0, len(emails), genuine


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------


def _load_env() -> None:
    env = PROJECT_ROOT / ".env"
    if env.exists():
        for line in env.read_text().splitlines():
            line = line.strip()
            if not line or line.startswith("#") or "=" not in line:
                continue
            k, v = line.split("=", 1)
            k, v = k.strip(), v.strip().strip('"').strip("'")
            if k == "API_SECRET_KEY":
                globals()["API_KEY"] = v
            elif k == "IRA_API_URL":
                globals()["API_URL"] = v.strip()


def main() -> None:
    import argparse
    _load_env()
    parser = argparse.ArgumentParser(description="Top 50 hottest leads from 07_Leads, excluding customers.")
    parser.add_argument("--output", "-o", type=Path, help="Write top 50 CSV here.")
    parser.add_argument("--dry-run", action="store_true", help="Only exclude customers; do not call API (no hottest ranking).")
    parser.add_argument("--limit", type=int, default=0, help="Limit leads to check (0 = all).")
    args = parser.parse_args()

    print("Loading customers (to exclude)...", flush=True)
    customer_emails, customer_companies = load_all_customers()
    print(f"  Customer emails: {len(customer_emails)}")
    print(f"  Customer companies (normalized): {len(customer_companies)}")

    print("Loading leads from 07_Leads_and_Contacts...", flush=True)
    leads = load_all_leads_07_only()
    print(f"  Total unique leads (by email): {len(leads)}")

    # Exclude customers
    def is_customer(lead: dict[str, str]) -> bool:
        if lead["email"] in customer_emails:
            return True
        cn = _normalize_company(lead.get("company") or "")
        if cn and cn in customer_companies:
            return True
        # Fuzzy: company name contained in any customer company or vice versa
        for c in customer_companies:
            if len(c) < 4:
                continue
            if c in cn or cn in c:
                return True
        return False

    leads_no_customers = [L for L in leads if not is_customer(L)]
    print(f"  After excluding customers: {len(leads_no_customers)}")

    if not leads_no_customers:
        print("No leads left after excluding customers.")
        return

    if args.dry_run:
        print("\n[DRY RUN] First 30 leads (email | company | source):")
        for L in leads_no_customers[:30]:
            print(f"  {L['email']} | {L['company']} | {L['source']}")
        if len(leads_no_customers) > 30:
            print(f"  ... and {len(leads_no_customers) - 30} more")
        return

    if not httpx:
        print("Install httpx to call Ira API: poetry add httpx", file=sys.stderr)
        return

    to_check = leads_no_customers if args.limit <= 0 else leads_no_customers[: args.limit]
    communicated: list[dict[str, Any]] = []

    for i, lead in enumerate(to_check, 1):
        email = lead["email"]
        print(f"[{i}/{len(to_check)}] {email} ({lead['company']})...", flush=True)
        ok, total, genuine = has_communicated(email)
        if ok:
            row = {**lead, "emails_from_them": total, "genuine_replies": genuine}
            communicated.append(row)

    # Hottest = most genuine replies, then most total from them
    communicated.sort(key=lambda r: (-r.get("genuine_replies", 0), -r.get("emails_from_them", 0)))
    top50 = communicated[:50]

    print(f"\n--- Summary ---")
    print(f"Checked: {len(to_check)}")
    print(f"Communicated (they replied or sent us email): {len(communicated)}")
    print(f"Top 50 hottest (by genuine replies, then total):")

    print("\n" + "=" * 100)
    for i, r in enumerate(top50, 1):
        print(f"{i:2}. {r['company'][:40]:40} | {r['email']:35} | genuine={r.get('genuine_replies', 0):2} from_them={r.get('emails_from_them', 0):2} | {r['source']}")
    print("=" * 100)

    if args.output and top50:
        args.output.parent.mkdir(parents=True, exist_ok=True)
        with open(args.output, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(
                f,
                fieldnames=["email", "name", "company", "source", "emails_from_them", "genuine_replies"],
            )
            w.writeheader()
            w.writerows(
                {k: r.get(k, "") for k in ["email", "name", "company", "source", "emails_from_them", "genuine_replies"]}
                for r in top50
            )
        print(f"\nWrote top 50 CSV: {args.output}")


if __name__ == "__main__":
    main()
