#!/usr/bin/env python3
"""
Build the candidate database (applicant datasheets) from fragmented files in data/imports.

Reads:
- data/imports/22_HR Data/Job Application (Responses).xlsx — Google Form responses
- data/imports/22_HR Data/CAD candidate.xlsx — CAD candidates with HR status
- data/imports/22_HR Data/Shortlisted candidates update.xlsx — sheet "list" (NAME, MAIL ID, DESIGNATION, STATUS)

Maps rows to canonical candidate profile and upserts into the SQLite candidate store
(data/brain/candidates.db). Run this before engaging applicants so Anu has full
datasheets. Optionally run after indexing/ingesting 22_HR Data so the KB has the
HR docs for context.

Usage:
  poetry run python scripts/build_candidate_database_from_imports.py
  poetry run python scripts/build_candidate_database_from_imports.py --dry-run
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from pathlib import Path

PROJECT_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(PROJECT_ROOT))

HR_DATA_DIR = PROJECT_ROOT / "data" / "imports" / "22_HR Data"


def _norm_email(v: object) -> str | None:
    if v is None:
        return None
    s = str(v).strip().lower()
    if "@" in s and "." in s:
        return s
    return None


def _norm_str(v: object) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _norm_float(v: object) -> float | None:
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


async def load_job_application_responses() -> list[dict]:
    """Load Job Application (Responses).xlsx and yield candidate-like dicts."""
    try:
        import openpyxl
    except ImportError:
        return []
    path = HR_DATA_DIR / "Job Application (Responses).xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = None
    out = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(row)]
            continue
        if not row:
            continue
        row_dict = dict(zip(headers, row))
        email = _norm_email(row_dict.get("Email Address") or row_dict.get("Email"))
        if not email:
            continue
        name = _norm_str(row_dict.get("Name"))
        phone = _norm_str(row_dict.get("Phone number"))
        position = _norm_str(row_dict.get("Which position(s) are you interested in?"))
        experience_years = _norm_float(row_dict.get("What is your total work experience? (in years)"))
        current_company = _norm_str(row_dict.get("Which company you work at currently?"))
        location = _norm_str(row_dict.get("Where are you currently located?"))
        strengths = _norm_str(row_dict.get("What are your strengths?"))
        softwares = _norm_str(row_dict.get("What softwares do you work with?"))
        why_hire = _norm_str(row_dict.get("Why should we hire you?"))
        profile = {
            "name": name or None,
            "email": email,
            "phone": phone or None,
            "current_role": position or None,
            "current_company": current_company or None,
            "location": location or None,
            "experience_years": experience_years,
            "skills": [s.strip() for s in softwares.replace(",", " ").split() if s.strip()] if softwares else [],
            "summary": (why_hire or "")[:500] if why_hire else None,
            "experience_highlights": [strengths[:300]] if strengths else [],
        }
        out.append({
            "email": email,
            "name": name or None,
            "profile": profile,
            "source_type": "file",
            "source_id": str(path.relative_to(PROJECT_ROOT)),
            "notes": "Job Application (Responses).xlsx",
        })
    wb.close()
    return out


async def load_cad_candidates() -> list[dict]:
    """Load CAD candidate.xlsx and yield candidate-like dicts."""
    try:
        import openpyxl
    except ImportError:
        return []
    path = HR_DATA_DIR / "CAD candidate.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    ws = wb.active
    headers = None
    out = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(row)]
            continue
        if not row:
            continue
        row_dict = dict(zip(headers, row))
        email = _norm_email(row_dict.get("Email"))
        if not email:
            continue
        name = _norm_str(row_dict.get("Full Name"))
        phone = _norm_str(row_dict.get("Phone number"))
        experience_years = _norm_float(row_dict.get("Years of Work Experience"))
        current_role = _norm_str(row_dict.get("Current Job Title"))
        status_hr = _norm_str(row_dict.get("Status from HR"))
        profile = {
            "name": name or None,
            "email": email,
            "phone": phone or None,
            "current_role": current_role or None,
            "experience_years": experience_years,
            "summary": status_hr[:800] if status_hr else None,
            "experience_highlights": [status_hr[:400]] if status_hr else [],
        }
        out.append({
            "email": email,
            "name": name or None,
            "profile": profile,
            "source_type": "file",
            "source_id": str(path.relative_to(PROJECT_ROOT)),
            "notes": status_hr[:500] if status_hr else None,
        })
    wb.close()
    return out


async def load_shortlisted_candidates() -> list[dict]:
    """Load Shortlisted candidates update.xlsx, sheet 'list' (SR NO, NAME, MAIL ID, PHONE NUMBER, DESIGNATION, STATUS)."""
    try:
        import openpyxl
    except ImportError:
        return []
    path = HR_DATA_DIR / "Shortlisted candidates update.xlsx"
    if not path.exists():
        return []
    wb = openpyxl.load_workbook(path, read_only=True)
    if "list" not in wb.sheetnames:
        wb.close()
        return []
    ws = wb["list"]
    headers = None
    out = []
    for row in ws.iter_rows(values_only=True):
        if headers is None:
            headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(row)]
            continue
        if not row:
            continue
        row_dict = dict(zip(headers, row))
        email = _norm_email(row_dict.get("MAIL ID") or row_dict.get("Mail ID") or row_dict.get("Email"))
        if not email:
            continue
        name = _norm_str(row_dict.get("NAME") or row_dict.get("Name"))
        phone = _norm_str(row_dict.get("PHONE NUMBER") or row_dict.get("Phone number"))
        current_role = _norm_str(row_dict.get("DESIGNATION ") or row_dict.get("DESIGNATION"))
        status = _norm_str(row_dict.get("STATUS"))
        profile = {
            "name": name or None,
            "email": email,
            "phone": phone or None,
            "current_role": current_role or None,
            "summary": status[:800] if status else None,
            "experience_highlights": [status[:400]] if status else [],
        }
        notes = f"Shortlisted list: {status}" if status else "Shortlisted candidates update.xlsx"
        out.append({
            "email": email,
            "name": name or None,
            "profile": profile,
            "source_type": "file",
            "source_id": str(path.relative_to(PROJECT_ROOT)) + "#list",
            "notes": notes,
        })
    wb.close()
    return out


async def main() -> int:
    parser = argparse.ArgumentParser(description="Build candidate DB from 22_HR Data imports")
    parser.add_argument("--dry-run", action="store_true", help="Print what would be upserted, do not write DB")
    args = parser.parse_args()

    if not HR_DATA_DIR.exists():
        print(f"HR Data folder not found: {HR_DATA_DIR}", file=sys.stderr)
        return 1

    from ira.data.candidates import CandidateStore

    job_app = await load_job_application_responses()
    cad = await load_cad_candidates()
    shortlisted = await load_shortlisted_candidates()
    by_email: dict[str, dict] = {}
    for c in job_app:
        by_email[c["email"]] = c
    for c in cad:
        if c["email"] not in by_email:
            by_email[c["email"]] = c
        else:
            existing = by_email[c["email"]]
            existing["profile"] = {**existing.get("profile", {}), **c.get("profile", {})}
            if c.get("notes"):
                existing["notes"] = (existing.get("notes") or "") + "\n" + c["notes"]
    for c in shortlisted:
        if c["email"] not in by_email:
            by_email[c["email"]] = c
        else:
            existing = by_email[c["email"]]
            existing["profile"] = {**existing.get("profile", {}), **c.get("profile", {})}
            if c.get("notes"):
                existing["notes"] = (existing.get("notes") or "") + "\n" + c["notes"]

    if args.dry_run:
        print(f"Would upsert {len(by_email)} candidates from Job Application + CAD candidate + Shortlisted sheets.")
        for email, c in list(by_email.items())[:10]:
            print(f"  {email} | {c.get('name')} | {c.get('source_id')}")
        if len(by_email) > 10:
            print(f"  ... and {len(by_email) - 10} more")
        return 0

    store = CandidateStore()
    for email, c in by_email.items():
        try:
            await store.upsert(
                email,
                name=c.get("name"),
                source_type=c.get("source_type"),
                source_id=c.get("source_id"),
                profile=c.get("profile"),
                notes=c.get("notes"),
            )
        except Exception as e:
            print(f"  Skip {email}: {e}", file=sys.stderr)
    n = await store.count()
    print(f"Candidate database built: {n} candidates in data/brain/candidates.db")
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
